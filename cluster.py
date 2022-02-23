import openpyxl
import pylab as pl
import numpy as np
from sklearn import preprocessing
from sklearn.decomposition import PCA
from sklearn import preprocessing
from sklearn.cluster import KMeans
import pandas as pd
import os
import joblib

pl.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
pl.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

data_book0 = openpyxl.load_workbook('0.xlsx')
data_book1 = openpyxl.load_workbook('1.xlsx')
data_book2 = openpyxl.load_workbook('2.xlsx')
data_book3 = openpyxl.load_workbook('3.xlsx')
row_list = []
X = []
Y = []
# 0普通 1好转 2重症 3死亡
for i in range(0, len(data_book0.sheetnames)):
    data_sheet = data_book0.worksheets[i]
    data_row = data_sheet.max_row
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=2, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    Y.append(0)
    row_list.clear()
for i in range(0, len(data_book1.sheetnames)):
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=2, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    Y.append(0)
    row_list.clear()
for i in range(0, len(data_book0.sheetnames)):
    data_sheet = data_book0.worksheets[i]
    data_row = data_sheet.max_row
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=data_row, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    row_list.clear()
    Y.append(1)
for i in range(0, len(data_book2.sheetnames)):
    data_sheet = data_book2.worksheets[i]
    data_row = data_sheet.max_row
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=data_row, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    row_list.clear()
    Y.append(1)
for i in range(0, len(data_book1.sheetnames)):
    data_sheet = data_book1.worksheets[i]
    data_row = data_sheet.max_row
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=data_row, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    row_list.clear()
    Y.append(2)
for i in range(0, len(data_book2.sheetnames)):
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=2, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    Y.append(2)
    row_list.clear()
for i in range(0, len(data_book3.sheetnames)):
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=2, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    Y.append(2)
    row_list.clear()
for i in range(0, len(data_book3.sheetnames)):
    data_sheet = data_book3.worksheets[i]
    data_row = data_sheet.max_row
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=data_row, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    row_list.clear()
    Y.append(3)

Y = np.array(Y)
# z-score标准化处理
zscore = preprocessing.StandardScaler()
data = zscore.fit_transform(X)

pca = PCA(n_components=9)
X_p = pca.fit(data).transform(data)
print(sum(pca.explained_variance_ratio_))
print(pca.explained_variance_ratio_)
print(pca.explained_variance_)

kmodel = KMeans(n_clusters=4, n_jobs=4)
kmodel.fit(X_p)
label = pd.Series(kmodel.labels_)  # 各样本的类别
print(label)

clustered_label = np.array(label)
print(clustered_label)

num = pd.Series(kmodel.labels_).value_counts()  # 统计各样本对应的类别的数目
center = pd.DataFrame(kmodel.cluster_centers_)  # 找出聚类中心
print(center)
print(Y)
right = 0
for i in range(len(Y)):
    if(Y[i] == clustered_label[i]):
        right = right + 1
print(right/len(Y))

dirs = 'testModel'
if not os.path.exists(dirs):
    os.makedirs(dirs)

# 保存模型
joblib.dump(kmodel, dirs + '/kmodel.pkl')