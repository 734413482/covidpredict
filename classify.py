import openpyxl
import pylab as pl
import numpy as np
from sklearn import preprocessing
from sklearn.decomposition import PCA
from sklearn import preprocessing
from sklearn.cluster import KMeans
import pandas as pd
import os
import joblib
import pylab as pl
import openpyxl
import numpy as np
from sklearn.ensemble import VotingClassifier, RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import MultinomialNB
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import accuracy_score, classification_report


pl.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
pl.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

data_book0 = openpyxl.load_workbook('0.xlsx')
data_book1 = openpyxl.load_workbook('1.xlsx')
data_book2 = openpyxl.load_workbook('2.xlsx')
data_book3 = openpyxl.load_workbook('3.xlsx')
row_list = []
X = []
Y = []
# 0普通 1好转 2重症 3死亡
for i in range(0, len(data_book0.sheetnames)):
    data_sheet = data_book0.worksheets[i]
    data_row = data_sheet.max_row
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=2, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    Y.append(0)
    row_list.clear()
for i in range(0, len(data_book1.sheetnames)):
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=2, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    Y.append(0)
    row_list.clear()
for i in range(0, len(data_book0.sheetnames)):
    data_sheet = data_book0.worksheets[i]
    data_row = data_sheet.max_row
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=data_row, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    row_list.clear()
    Y.append(1)
for i in range(0, len(data_book2.sheetnames)):
    data_sheet = data_book2.worksheets[i]
    data_row = data_sheet.max_row
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=data_row, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    row_list.clear()
    Y.append(1)
for i in range(0, len(data_book1.sheetnames)):
    data_sheet = data_book1.worksheets[i]
    data_row = data_sheet.max_row
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=data_row, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    row_list.clear()
    Y.append(2)
for i in range(0, len(data_book2.sheetnames)):
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=2, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    Y.append(2)
    row_list.clear()
for i in range(0, len(data_book3.sheetnames)):
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=2, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    Y.append(2)
    row_list.clear()
for i in range(0, len(data_book3.sheetnames)):
    data_sheet = data_book3.worksheets[i]
    data_row = data_sheet.max_row
    for k in range(2, 22):
        row_list.append(data_sheet.cell(row=data_row, column=k).value)
    temp_mat_list = list(row_list)
    X.append(temp_mat_list)
    row_list.clear()
    Y.append(3)

Y = np.array(Y)
# z-score标准化处理
zscore = preprocessing.StandardScaler()
data = zscore.fit_transform(X)

pca = PCA(n_components=9)
X_p = pca.fit(data).transform(data)
print(sum(pca.explained_variance_ratio_))
print(pca.explained_variance_ratio_)
print(pca.explained_variance_)

print(type(X_p))

X_train, X_test, Y_train, Y_test = train_test_split(X_p, Y, test_size=0.2, random_state=42)
DTC_clf = DecisionTreeClassifier(max_depth=5,
                             min_weight_fraction_leaf=0.015,
                             min_samples_leaf=4,
                             min_samples_split=5,
                             max_leaf_nodes=5,
                             random_state=42
                             )
KNNuni_clf = KNeighborsClassifier(n_neighbors=6, weights="uniform", p=1)
KNNdis_clf = KNeighborsClassifier(n_neighbors=6, weights="distance", p=1)
voting_clf = VotingClassifier(
    estimators=[('dtc', DTC_clf), ('knnu', KNNuni_clf), ('knnd', KNNdis_clf)],
    voting='soft'
)

DTC_clf.fit(X_train, Y_train)
KNNuni_clf.fit(X_train, Y_train)
KNNdis_clf.fit(X_train, Y_train)
voting_clf.fit(X_train, Y_train)

for clf in (DTC_clf, KNNuni_clf, KNNdis_clf, voting_clf):
    clf.fit(X_train, Y_train)
    y_pred = clf.predict(X_test)
    print(clf.__class__.__name__, accuracy_score(Y_test, y_pred))