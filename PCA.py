import openpyxl
import pylab as pl
import numpy as np
from sklearn import preprocessing
from sklearn.decomposition import PCA

pl.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
pl.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

data_book0 = openpyxl.load_workbook('0.xlsx')
data_book1 = openpyxl.load_workbook('1.xlsx')
data_book2 = openpyxl.load_workbook('2.xlsx')
data_book3 = openpyxl.load_workbook('3.xlsx')
row_list = []
X = []
Y = []

for i in range(0, len(data_book0.sheetnames)):
    data_sheet = data_book0.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row + 1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_mat_list = list(row_list)
        X.append(temp_mat_list)
        row_list.clear()
for i in range(0, len(data_book1.sheetnames)):
    data_sheet = data_book1.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row + 1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_mat_list = list(row_list)
        X.append(temp_mat_list)
        row_list.clear()
for i in range(0, len(data_book2.sheetnames)):
    data_sheet = data_book2.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row + 1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_mat_list = list(row_list)
        X.append(temp_mat_list)
        row_list.clear()
for i in range(0, len(data_book3.sheetnames)):
    data_sheet = data_book3.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row + 1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_mat_list = list(row_list)
        X.append(temp_mat_list)
        row_list.clear()
X = np.array(X)
print(X.shape)
# z-score标准化处理
zscore = preprocessing.StandardScaler()
data = zscore.fit_transform(X)

pca = PCA(n_components=9)
X_p = pca.fit(data).transform(data)
print(sum(pca.explained_variance_ratio_))
print(pca.explained_variance_ratio_)
print(pca.explained_variance_)