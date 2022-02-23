import openpyxl
import pylab as pl
from sklearn import preprocessing
from sklearn.cluster import KMeans
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

pl.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
pl.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

data_book0 = openpyxl.load_workbook('0.xlsx')
data_book1 = openpyxl.load_workbook('1.xlsx')
data_book2 = openpyxl.load_workbook('2.xlsx')
data_book3 = openpyxl.load_workbook('3.xlsx')
row_list = []
X = []
Y = []
for i in range(0, len(data_book0.sheetnames)):
    data_sheet = data_book0.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row + 1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_mat_list = list(row_list)
        X.append(temp_mat_list)
        row_list.clear()
for i in range(0, len(data_book1.sheetnames)):
    data_sheet = data_book1.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row + 1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_mat_list = list(row_list)
        X.append(temp_mat_list)
        row_list.clear()
for i in range(0, len(data_book2.sheetnames)):
    data_sheet = data_book2.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row + 1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_mat_list = list(row_list)
        X.append(temp_mat_list)
        row_list.clear()
for i in range(0, len(data_book3.sheetnames)):
    data_sheet = data_book3.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row + 1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_mat_list = list(row_list)
        X.append(temp_mat_list)
        row_list.clear()

# z-score标准化处理
# zscore = preprocessing.StandardScaler()
# data = zscore.fit_transform(X)

# maxmin标准化处理
minmax = preprocessing.MinMaxScaler()
data = minmax.fit_transform(X)
print(data)

kmodel = KMeans(n_clusters=10, n_jobs=4)
kmodel.fit(data)
label = pd.Series(kmodel.labels_)  # 各样本的类别
print(label)
num = pd.Series(kmodel.labels_).value_counts()  # 统计各样本对应的类别的数目
center = pd.DataFrame(kmodel.cluster_centers_)  # 找出聚类中心
print(center)
