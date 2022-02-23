import openpyxl

data_book = openpyxl.load_workbook('重症-死亡.xlsx')

def get_median(data):
    data.sort()
    half = len(data) // 2
    return (data[half] + data[~half]) / 2
col = 21
for i in range(0, len(data_book.sheetnames)):#
    data_sheet = data_book.worksheets[i]
    data_row = data_sheet.max_row
    data_feature = []
    for j in range(2, data_row + 1):
        if(data_sheet.cell(row=j, column=col).value != 0):
            data_feature.append(data_sheet.cell(row=j, column=col).value)
    print(data_feature)
    median = get_median(data_feature)
    print(median)
    for j in range(2, data_row + 1):
        if (data_sheet.cell(row=j, column=col).value == 0):
            data_sheet.cell(row=j, column=col).value = median
col = 20
for i in range(0, len(data_book.sheetnames)):#
    data_sheet = data_book.worksheets[i]
    data_row = data_sheet.max_row
    data_feature = []
    for j in range(2, data_row + 1):
        if(data_sheet.cell(row=j, column=col).value != 0):
            data_feature.append(data_sheet.cell(row=j, column=col).value)
    print(data_feature)
    median = get_median(data_feature)
    print(median)
    for j in range(2, data_row + 1):
        if (data_sheet.cell(row=j, column=col).value == 0):
            data_sheet.cell(row=j, column=col).value = median
data_book.save('3.xlsx')

